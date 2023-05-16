import sys
import openpyxl
from openpyxl.styles import PatternFill

def highlight_cells(file_name, text):
    # Open the spreadsheet
    workbook = openpyxl.load_workbook(file_name)

    # Highlight pattern
    highlight = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Iterate over all sheets and rows/columns
    for sheet in workbook.sheetnames:
        for row in workbook[sheet].iter_rows():
            for cell in row:
                # If the cell contains the text, apply the highlight
                if text in str(cell.value):
                    cell.fill = highlight

    # Save the workbook
    workbook.save(file_name)

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python highlight_cells.py [file_name] [text]")
    else:
        highlight_cells(sys.argv[1], sys.argv[2])
